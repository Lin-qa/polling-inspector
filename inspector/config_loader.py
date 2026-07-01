from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from inspector.models import CheckItem, InspectorConfig, NotifyGroup, PreRequest

CHECK_SHEET = "巡检接口"
VARIABLE_SHEET = "前置变量"
NOTIFY_SHEET = "通知配置"
PRE_REQUEST_SHEET = "前置请求"


def create_template(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheets = [
        (
            CHECK_SHEET,
            [
                "是否启用",
                "场景名称",
                "接口名称",
                "请求方式",
                "URL",
                "请求头JSON",
                "请求参数",
                "成功判断",
                "轮询间隔秒",
                "异常后轮询间隔秒",
                "超时时间ms",
                "通知组",
                "前置请求",
            ],
            [
                [
                    "是",
                    "示例巡检",
                    "健康检查",
                    "GET",
                    "https://api.example.test/health",
                    "{}",
                    "无",
                    "status=200",
                    3600,
                    600,
                    5000,
                    "默认组",
                    "",
                ],
                [
                    "否",
                    "示例巡检",
                    "需要登录的业务状态",
                    "POST",
                    "https://api.example.test/status",
                    '{"Content-Type":"application/json"}',
                    '{"ticketid":"${ticketid}"}',
                    "status=200; code=0",
                    3600,
                    600,
                    5000,
                    "默认组",
                    "登录刷新ticket",
                ],
                [
                    "否",
                    "示例巡检",
                    "派生参数业务查询",
                    "POST",
                    "https://api.example.test/member/coupon",
                    '{"Content-Type":"application/json"}',
                    '{"memberId":"${member_id}"}',
                    "status=200; code=0",
                    3600,
                    600,
                    5000,
                    "默认组",
                    "登录刷新ticket,会员详情取memberId",
                ],
            ],
        ),
        (
            PRE_REQUEST_SHEET,
            [
                "是否启用",
                "前置名称",
                "请求方式",
                "URL",
                "请求头JSON",
                "请求参数",
                "成功判断",
                "提取变量JSON",
                "超时时间ms",
            ],
            [
                [
                    "是",
                    "登录刷新ticket",
                    "POST",
                    "https://api.example.test/login",
                    '{"Content-Type":"application/json"}',
                    '{"openid":"${openid}","unionid":"${unionid}"}',
                    "status=200; code=0",
                    '{"ticketid":"data.ticketid","app_user_id":"data.useruuid","phone":"data.phone"}',
                    5000,
                ],
                [
                    "是",
                    "会员详情取memberId",
                    "POST",
                    "https://api.example.test/member/detail",
                    '{"Content-Type":"application/json"}',
                    '{"appId":"${app_user_id}"}',
                    "status=200; code=0",
                    '{"member_id":"data.memId"}',
                    5000,
                ],
            ],
        ),
        (
            VARIABLE_SHEET,
            ["变量名", "变量值", "是否敏感", "说明"],
            [
                ["openid", "demo-openid", "是", "登录前置请求入参，真实 openid 不要提交到 git"],
                ["unionid", "demo-unionid", "是", "登录前置请求入参，真实 unionid 不要提交到 git"],
            ],
        ),
        (
            NOTIFY_SHEET,
            ["通知组", "Webhook地址", "通知类型", "签名密钥", "是否@所有人", "备注"],
            [
                ["默认组", "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=REPLACE_WITH_YOUR_KEY", "企业微信", "", "否", "企业微信或飞书群机器人 webhook"],
            ],
        ),
    ]

    for index, (title, headers, rows) in enumerate(sheets):
        ws = wb.active if index == 0 else wb.create_sheet(title)
        ws.title = title
        ws.append(headers)
        for row in rows:
            ws.append(row)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = center
        widths = _widths_for_sheet(title)
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"

    wb.save(path)


def load_config(path: Path) -> InspectorConfig:
    wb = load_workbook(path, data_only=True)
    missing = [name for name in [CHECK_SHEET, VARIABLE_SHEET, NOTIFY_SHEET] if name not in wb.sheetnames]
    if missing:
        raise ValueError(f"配置文件缺少 sheet：{', '.join(missing)}")

    variables, sensitive_variables = _load_variables(wb[VARIABLE_SHEET])
    notify_groups = _load_notify_groups(wb[NOTIFY_SHEET])
    pre_requests = _load_pre_requests(wb[PRE_REQUEST_SHEET]) if PRE_REQUEST_SHEET in wb.sheetnames else {}
    checks = _load_checks(wb[CHECK_SHEET])
    enabled_checks = [item for item in checks if item.enabled]
    if not enabled_checks:
        raise ValueError("没有启用的巡检接口，请检查「巡检接口」sheet 的「是否启用」列。")
    return InspectorConfig(
        checks=enabled_checks,
        variables=variables,
        sensitive_variables=sensitive_variables,
        notify_groups=notify_groups,
        pre_requests=pre_requests,
    )


def _load_checks(sheet) -> list[CheckItem]:
    header_index = _header_index(sheet)
    checks: list[CheckItem] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        checks.append(
            CheckItem(
                enabled=_enabled(_value(row, header_index, "是否启用"), default=True),
                scenario_name=str(_value(row, header_index, "场景名称") or "").strip(),
                api_name=str(_value(row, header_index, "接口名称") or "").strip(),
                method=str(_value(row, header_index, "请求方式") or "GET").upper().strip(),
                url=str(_value(row, header_index, "URL") or "").strip(),
                headers=_parse_json_object(_value(row, header_index, "请求头JSON"), "请求头JSON"),
                params=str(_value(row, header_index, "请求参数") or "").strip(),
                success_rule=str(_value(row, header_index, "成功判断") or "").strip(),
                interval_seconds=_float_or_default(
                    _value(row, header_index, "轮询间隔秒"),
                    _legacy_ms_to_seconds(_value_any(row, header_index, ["轮询间隔ms", "轮询间隔毫秒"]), 3600),
                ),
                abnormal_interval_seconds=_float_or_default(
                    _value_any(row, header_index, ["异常后轮询间隔秒", "异常轮询间隔秒", "异常后轮询时间秒"]),
                    600,
                ),
                timeout_ms=_int_or_default(_value_any(row, header_index, ["超时时间ms", "超时时间毫秒"]), _legacy_seconds_to_ms(_value(row, header_index, "超时时间秒"), 5000)),
                notify_group=str(_value(row, header_index, "通知组") or "默认组").strip(),
                pre_request_name=str(_value(row, header_index, "前置请求") or "").strip(),
            )
        )
    for item in checks:
        if item.enabled and (not item.scenario_name or not item.api_name or not item.url):
            raise ValueError("启用的巡检接口必须填写场景名称、接口名称和 URL。")
    return checks


def _load_pre_requests(sheet) -> dict[str, PreRequest]:
    header_index = _header_index(sheet)
    pre_requests: dict[str, PreRequest] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        if not _enabled(_value(row, header_index, "是否启用"), default=True):
            continue
        name = str(_value(row, header_index, "前置名称") or "").strip()
        if not name:
            raise ValueError("启用的前置请求必须填写前置名称。")
        pre_requests[name] = PreRequest(
            name=name,
            method=str(_value(row, header_index, "请求方式") or "GET").upper().strip(),
            url=str(_value(row, header_index, "URL") or "").strip(),
            headers=_parse_json_object(_value(row, header_index, "请求头JSON"), "请求头JSON"),
            params=str(_value(row, header_index, "请求参数") or "").strip(),
            success_rule=str(_value(row, header_index, "成功判断") or "").strip(),
            extractors=_parse_json_object(_value(row, header_index, "提取变量JSON"), "提取变量JSON"),
            timeout_ms=_int_or_default(_value_any(row, header_index, ["超时时间ms", "超时时间毫秒"]), 5000),
        )
    for item in pre_requests.values():
        if not item.url:
            raise ValueError(f"前置请求「{item.name}」必须填写 URL。")
    return pre_requests


def _load_variables(sheet) -> tuple[dict[str, str], set[str]]:
    header_index = _header_index(sheet)
    variables: dict[str, str] = {}
    sensitive: set[str] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        key = str(_value(row, header_index, "变量名") or "").strip()
        if not key:
            continue
        variables[key] = str(_value(row, header_index, "变量值") or "")
        if _enabled(_value(row, header_index, "是否敏感"), default=False):
            sensitive.add(key)
    return variables, sensitive


def _load_notify_groups(sheet) -> dict[str, NotifyGroup]:
    header_index = _header_index(sheet)
    groups: dict[str, NotifyGroup] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = str(_value(row, header_index, "通知组") or "").strip()
        if not name:
            continue
        groups[name] = NotifyGroup(
            name=name,
            webhook_url=str(_value_any(row, header_index, ["Webhook地址", "企业微信Webhook", "飞书Webhook"]) or "").strip(),
            webhook_type=str(_value_any(row, header_index, ["通知类型", "Webhook类型", "机器人类型"]) or "").strip(),
            secret=str(_value_any(row, header_index, ["签名密钥", "飞书签名密钥", "Secret"]) or "").strip(),
            mention_all=_enabled(_value(row, header_index, "是否@所有人"), default=False),
        )
    return groups


def _parse_json_object(value: Any, column_name: str) -> dict[str, str]:
    text = str(value or "").strip()
    if not text or text in {"无", "空", "{}"}:
        return {}
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ValueError(f"{column_name} 不是合法 JSON：{text}") from exc
    if not isinstance(data, dict):
        raise ValueError(f"{column_name} 必须是 JSON 对象。")
    return {str(key): str(val) for key, val in data.items()}


def _header_index(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): index
        for index, cell in enumerate(sheet[1])
        if cell.value is not None
    }


def _value(row: tuple, header_index: dict[str, int], header: str):
    index = header_index.get(header)
    if index is None or index >= len(row):
        return None
    return row[index]


def _value_any(row: tuple, header_index: dict[str, int], headers: list[str]):
    for header in headers:
        value = _value(row, header_index, header)
        if value not in (None, ""):
            return value
    return None


def _enabled(value: Any, default: bool) -> bool:
    text = str(value or "").strip()
    if not text:
        return default
    return text in {"是", "Y", "y", "yes", "YES", "true", "True", "1", "启用"}


def _float_or_default(value: Any, default: float) -> float:
    number = _float_or_none(value)
    return default if number is None else number


def _float_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _int_or_default(value: Any, default: int) -> int:
    if value in (None, ""):
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _legacy_seconds_to_ms(value: Any, default_ms: int) -> int:
    if value in (None, ""):
        return default_ms
    try:
        return int(float(value) * 1000)
    except (TypeError, ValueError):
        return default_ms


def _legacy_ms_to_seconds(value: Any, default_seconds: float) -> float:
    if value in (None, ""):
        return default_seconds
    try:
        return float(value) / 1000
    except (TypeError, ValueError):
        return default_seconds


def _widths_for_sheet(title: str) -> list[int]:
    if title == CHECK_SHEET:
        return [10, 18, 22, 10, 44, 38, 38, 24, 14, 18, 12, 14, 18]
    if title == PRE_REQUEST_SHEET:
        return [10, 18, 10, 44, 38, 38, 24, 36, 12]
    if title == VARIABLE_SHEET:
        return [18, 44, 12, 34]
    return [18, 70, 14, 34, 14, 34]
