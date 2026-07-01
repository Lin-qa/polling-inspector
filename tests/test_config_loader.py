from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook

from inspector.config_loader import load_config


class ConfigLoaderTests(unittest.TestCase):
    def test_loads_interval_seconds_and_timeout_ms(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "config.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "巡检接口"
            ws.append(["是否启用", "场景名称", "接口名称", "请求方式", "URL", "请求头JSON", "请求参数", "成功判断", "轮询间隔秒", "异常后轮询间隔秒", "超时时间ms", "通知组"])
            ws.append(["是", "场景", "接口", "GET", "https://api.example.test/health", "{}", "无", "status=200", 12, 34, 6789, "默认组"])
            ws = wb.create_sheet("前置变量")
            ws.append(["变量名", "变量值", "是否敏感", "说明"])
            ws = wb.create_sheet("通知配置")
            ws.append(["通知组", "企业微信Webhook", "是否@所有人", "备注"])
            ws.append(["默认组", "", "否", ""])
            wb.save(path)

            config = load_config(path)

        self.assertEqual(config.checks[0].interval_seconds, 12)
        self.assertEqual(config.checks[0].abnormal_interval_seconds, 34)
        self.assertEqual(config.checks[0].timeout_ms, 6789)

    def test_loads_feishu_notify_group_with_new_columns(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "config.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "巡检接口"
            ws.append(["是否启用", "场景名称", "接口名称", "请求方式", "URL", "请求头JSON", "请求参数", "成功判断", "轮询间隔秒", "异常后轮询间隔秒", "超时时间ms", "通知组"])
            ws.append(["是", "场景", "接口", "GET", "https://api.example.test/health", "{}", "无", "status=200", 12, 34, 6789, "默认组"])
            ws = wb.create_sheet("前置变量")
            ws.append(["变量名", "变量值", "是否敏感", "说明"])
            ws = wb.create_sheet("通知配置")
            ws.append(["通知组", "Webhook地址", "通知类型", "签名密钥", "是否@所有人", "备注"])
            ws.append(["默认组", "https://open.feishu.cn/open-apis/bot/v2/hook/demo", "飞书", "demo-secret", "否", ""])
            wb.save(path)

            config = load_config(path)

        group = config.notify_groups["默认组"]
        self.assertEqual(group.webhook_type, "飞书")
        self.assertEqual(group.secret, "demo-secret")
        self.assertIn("open.feishu.cn", group.webhook_url)

    def test_loads_pre_request_and_check_reference(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "config.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "巡检接口"
            ws.append(["是否启用", "场景名称", "接口名称", "请求方式", "URL", "请求头JSON", "请求参数", "成功判断", "轮询间隔秒", "异常后轮询间隔秒", "超时时间ms", "通知组", "前置请求"])
            ws.append(["是", "场景", "接口", "GET", "https://api.example.test/health", "{}", "无", "status=200", 12, 34, 6789, "默认组", "登录"])
            ws = wb.create_sheet("前置变量")
            ws.append(["变量名", "变量值", "是否敏感", "说明"])
            ws = wb.create_sheet("前置请求")
            ws.append(["是否启用", "前置名称", "请求方式", "URL", "请求头JSON", "请求参数", "成功判断", "提取变量JSON", "超时时间ms"])
            ws.append(["是", "登录", "POST", "https://api.example.test/login", "{}", '{"openid":"${openid}"}', "status=200", '{"ticketid":"result.ticketid"}', 5000])
            ws = wb.create_sheet("通知配置")
            ws.append(["通知组", "Webhook地址", "通知类型", "签名密钥", "是否@所有人", "备注"])
            ws.append(["默认组", "", "", "", "否", ""])
            wb.save(path)

            config = load_config(path)

        self.assertEqual(config.checks[0].pre_request_name, "登录")
        self.assertIn("登录", config.pre_requests)
        self.assertEqual(config.pre_requests["登录"].extractors["ticketid"], "result.ticketid")


if __name__ == "__main__":
    unittest.main()
